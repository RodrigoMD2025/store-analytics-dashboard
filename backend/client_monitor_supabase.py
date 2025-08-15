import os
import json
from datetime import datetime, timedelta, time
from zoneinfo import ZoneInfo
import logging
import pandas as pd
import requests
from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import matplotlib.pyplot as plt
import io
from openpyxl.drawing.image import Image as Img
import time as time_module
from supabase import create_client, Client
import hashlib
from dotenv import load_dotenv

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

# Configuração do log
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("log_extracao.log", encoding="utf-8")
    ]
)

# Configurações do Supabase - Usar variáveis de ambiente
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

# Configurações do Telegram - Usar variáveis de ambiente
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID")

def init_supabase():
    """Inicializa o cliente Supabase"""
    try:
        if not SUPABASE_URL or not SUPABASE_KEY:
            logging.error("Variáveis SUPABASE_URL e SUPABASE_KEY devem estar configuradas")
            return None
        
        supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
        logging.info("Cliente Supabase inicializado com sucesso")
        return supabase
    except Exception as e:
        logging.error(f"Erro ao inicializar Supabase: {e}")
        return None

def criar_tabelas_supabase():
    """
    Cria as tabelas necessárias no Supabase (apenas para referência)
    Execute estes comandos SQL no editor do Supabase:
    
    -- Tabela de execuções (sessões de coleta)
    CREATE TABLE IF NOT EXISTS execucoes (
        id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
        cliente_id INTEGER REFERENCES clientes(id),
        cliente_nome TEXT NOT NULL,
        total_lojas INTEGER DEFAULT 0,
        lojas_sincronizadas INTEGER DEFAULT 0,
        lojas_atrasadas INTEGER DEFAULT 0,
        percentual_sincronizadas NUMERIC(5,2) DEFAULT 0,
        percentual_atrasadas NUMERIC(5,2) DEFAULT 0,
        status TEXT DEFAULT 'processando', -- processando, sucesso, erro
        erro_detalhes TEXT,
        executado_em TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
        origem TEXT DEFAULT 'local', -- local, github_actions
        created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
    );

    -- Tabela de dados das lojas (histórico detalhado)
    CREATE TABLE IF NOT EXISTS lojas_dados (
        id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
        execucao_id UUID REFERENCES execucoes(id) ON DELETE CASCADE,
        cliente_id INTEGER REFERENCES clientes(id),
        cliente_nome TEXT NOT NULL,
        loja_nome TEXT NOT NULL,
        identificador TEXT NOT NULL,
        atualizado_em TIMESTAMP WITH TIME ZONE,
        sincronizada BOOLEAN DEFAULT FALSE,
        tempo_atraso_horas NUMERIC(10,2) DEFAULT 0,
        tempo_atraso_dias INTEGER DEFAULT 0,
        hash_loja TEXT, -- hash único para identificar loja
        data_coleta TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
        created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW()
    );

    -- Tabela de métricas agregadas por período
    CREATE TABLE IF NOT EXISTS metricas_periodicas (
        id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
        cliente_id INTEGER REFERENCES clientes(id),
        cliente_nome TEXT NOT NULL,
        data_referencia DATE NOT NULL,
        periodo TEXT NOT NULL, -- diario, semanal, mensal
        total_lojas INTEGER DEFAULT 0,
        lojas_sincronizadas INTEGER DEFAULT 0,
        lojas_atrasadas INTEGER DEFAULT 0,
        percentual_sincronizadas NUMERIC(5,2) DEFAULT 0,
        tempo_medio_atraso_horas NUMERIC(10,2) DEFAULT 0,
        maior_atraso_horas NUMERIC(10,2) DEFAULT 0,
        execucoes_periodo INTEGER DEFAULT 0,
        created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
        updated_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(),
        UNIQUE(cliente_nome, data_referencia, periodo)
    );

    -- Índices para performance
    CREATE INDEX IF NOT EXISTS idx_execucoes_cliente_data ON execucoes(cliente_nome, executado_em);
    CREATE INDEX IF NOT EXISTS idx_lojas_dados_cliente_data ON lojas_dados(cliente_nome, data_coleta);
    CREATE INDEX IF NOT EXISTS idx_lojas_dados_execucao ON lojas_dados(execucao_id);
    CREATE INDEX IF NOT EXISTS idx_metricas_cliente_periodo ON metricas_periodicas(cliente_nome, periodo, data_referencia);
    """
    pass

def gerar_hash_loja(cliente_nome, loja_nome, identificador):
    """Gera hash único para identificar uma loja específica"""
    data = f"{cliente_nome}_{loja_nome}_{identificador}"
    return hashlib.md5(data.encode('utf-8')).hexdigest()

def carregar_base_clientes():
    """Carrega a base de clientes do Supabase"""
    try:
        supabase = init_supabase()
        if not supabase:
            logging.error("Falha ao conectar com Supabase")
            return []
        
        response = supabase.table('clientes').select('*').execute()
        
        if response.data:
            logging.info(f"Base de clientes carregada com {len(response.data)} clientes")
            return response.data
        else:
            logging.warning("Nenhum cliente encontrado na base")
            return []
            
    except Exception as e:
        logging.error(f"Erro ao carregar base de clientes do Supabase: {e}")
        return carregar_base_clientes_local()

def carregar_base_clientes_local():
    """Fallback: carrega base de clientes do arquivo JSON local"""
    try:
        if not os.path.exists("data.json"):
            logging.warning("Arquivo data.json não encontrado para fallback")
            return []
        
        with open("data.json", "r", encoding="utf-8") as file:
            clientes_data = json.load(file)
        
        if "clientes" not in clientes_data:
            logging.error("Estrutura inválida no data.json")
            return []
        
        logging.info(f"Fallback: Base de clientes local carregada com {len(clientes_data['clientes'])} clientes")
        return clientes_data['clientes']
        
    except Exception as e:
        logging.error(f"Erro no fallback para arquivo local: {e}")
        return []

def criar_execucao(supabase, cliente_info):
    """Cria um registro de execução no Supabase"""
    try:
        execucao_data = {
            "cliente_id": cliente_info.get('id'),
            "cliente_nome": cliente_info.get('nome'),
            "status": "processando",
            "executado_em": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
            "origem": "github_actions" if os.getenv("GITHUB_ACTIONS") else "local"
        }
        
        response = supabase.table('execucoes').insert(execucao_data).execute()
        
        if response.data:
            execucao_id = response.data[0]['id']
            logging.info(f"Execução criada com ID: {execucao_id}")
            return execucao_id
        else:
            logging.error("Erro ao criar execução")
            return None
            
    except Exception as e:
        logging.error(f"Erro ao criar execução: {e}")
        return None

def finalizar_execucao(supabase, execucao_id, resumo, status="sucesso", erro_detalhes=""):
    """Finaliza uma execução com os dados coletados"""
    try:
        update_data = {
            "total_lojas": resumo.get('total', 0),
            "lojas_sincronizadas": resumo.get('sincronizadas', 0),
            "lojas_atrasadas": resumo.get('atrasadas', 0),
            "percentual_sincronizadas": round(resumo.get('percentual_sincronizadas', 0), 2),
            "percentual_atrasadas": round(resumo.get('percentual_atrasadas', 0), 2),
            "status": status,
            "erro_detalhes": erro_detalhes
        }
        
        response = supabase.table('execucoes').update(update_data).eq('id', execucao_id).execute()
        
        if response.data:
            logging.info(f"Execução {execucao_id} finalizada com status: {status}")
            return True
        else:
            logging.error(f"Erro ao finalizar execução {execucao_id}")
            return False
            
    except Exception as e:
        logging.error(f"Erro ao finalizar execução {execucao_id}: {e}")
        return False

def salvar_dados_lojas_supabase(supabase, execucao_id, df, cliente_info):
    """Salva dados detalhados das lojas no Supabase"""
    try:
        if df.empty:
            logging.info("DataFrame vazio, nada para salvar")
            return True
        
        dados_lojas = []
        
        for _, row in df.iterrows():
            # Calcular tempo de atraso em horas
            tempo_atraso_horas = 0
            tempo_atraso_dias = 0
            
            if row['Tempo Atraso'] != timedelta(0):
                tempo_atraso_horas = float(round(row['Tempo Atraso'].total_seconds() / 3600, 2))
                tempo_atraso_dias = int(row['Tempo Atraso'].days)
            
            loja_data = {
                "execucao_id": execucao_id,
                "cliente_id": cliente_info.get('id'),
                "cliente_nome": cliente_info.get('nome'),
                "loja_nome": row['Loja'],
                "identificador": row['Identificador'],
                "atualizado_em": row['Data Atualizacao'].isoformat() if pd.notna(row['Data Atualizacao']) else None,
                "sincronizada": bool(row['Sincronizada']),
                "tempo_atraso_horas": tempo_atraso_horas,
                "tempo_atraso_dias": tempo_atraso_dias,
                "hash_loja": gerar_hash_loja(cliente_info.get('nome'), row['Loja'], row['Identificador']),
                "data_coleta": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat()
            }
            
            dados_lojas.append(loja_data)
        
        # Inserir dados em lotes para melhor performance
        batch_size = 50
        total_inseridos = 0
        
        for i in range(0, len(dados_lojas), batch_size):
            batch = dados_lojas[i:i + batch_size]
            
            response = supabase.table('lojas_dados').insert(batch).execute()
            
            if response.data:
                total_inseridos += len(response.data)
            else:
                logging.error(f"Erro ao inserir lote {i//batch_size + 1}")
        
        logging.info(f"Total de {total_inseridos} registros de lojas inseridos no Supabase")
        return True
        
    except Exception as e:
        logging.error(f"Erro ao salvar dados das lojas no Supabase: {e}")
        return False

def atualizar_metricas_periodicas(supabase, cliente_info, resumo):
    """Atualiza métricas agregadas por período"""
    try:
        hoje = datetime.now(ZoneInfo("America/Sao_Paulo")).date()
        cliente_nome = cliente_info.get('nome')
        
        # Calcular tempo médio de atraso (seria necessário dados mais detalhados)
        tempo_medio_atraso = 0  # Simplificado por agora
        maior_atraso = 0  # Simplificado por agora
        
        # Dados das métricas diárias
        metrica_data = {
            "cliente_id": cliente_info.get('id'),
            "cliente_nome": cliente_nome,
            "data_referencia": hoje.isoformat(),
            "periodo": "diario",
            "total_lojas": resumo.get('total', 0),
            "lojas_sincronizadas": resumo.get('sincronizadas', 0),
            "lojas_atrasadas": resumo.get('atrasadas', 0),
            "percentual_sincronizadas": round(resumo.get('percentual_sincronizadas', 0), 2),
            "tempo_medio_atraso_horas": tempo_medio_atraso,
            "maior_atraso_horas": maior_atraso,
            "execucoes_periodo": 1,
            "updated_at": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat()
        }
        
        # Tentar atualizar registro existente ou inserir novo
        existing = supabase.table('metricas_periodicas').select('*').eq('cliente_nome', cliente_nome).eq('data_referencia', hoje.isoformat()).eq('periodo', 'diario').execute()
        
        if existing.data:
            # Atualizar registro existente
            current_data = existing.data[0]
            metrica_data['execucoes_periodo'] = current_data.get('execucoes_periodo', 0) + 1
            
            response = supabase.table('metricas_periodicas').update(metrica_data).eq('id', current_data['id']).execute()
        else:
            # Inserir novo registro
            response = supabase.table('metricas_periodicas').insert(metrica_data).execute()
        
        if response.data:
            logging.info(f"Métricas periódicas atualizadas para {cliente_nome}")
            return True
        else:
            logging.error(f"Erro ao atualizar métricas periódicas para {cliente_nome}")
            return False
            
    except Exception as e:
        logging.error(f"Erro ao atualizar métricas periódicas: {e}")
        return False

def log_execucao(cliente_nome, status, detalhes="", total_lojas=0):
    """Mantém compatibilidade com logs existentes"""
    try:
        supabase = init_supabase()
        if not supabase:
            return
        
        log_data = {
            "cliente_nome": cliente_nome,
            "status": status,
            "detalhes": detalhes,
            "total_lojas": total_lojas,
            "executado_em": datetime.now(ZoneInfo("America/Sao_Paulo")).isoformat(),
            "origem": "github_actions" if os.getenv("GITHUB_ACTIONS") else "local"
        }
        
        response = supabase.table('logs_execucao').insert(log_data).execute()
        logging.info(f"Log registrado para {cliente_nome}: {status}")
        
    except Exception as e:
        logging.error(f"Erro ao registrar log no Supabase: {e}")

def setup_browser(headless: bool = True):
    """Configura o navegador Playwright"""
    try:
        playwright = sync_playwright().start()
        
        launch_options = {
            "headless": headless,
            "args": [
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-accelerated-2d-canvas",
                "--no-first-run",
                "--no-zygote",
                "--disable-gpu"
            ] if os.getenv("GITHUB_ACTIONS") else []
        }
        
        browser = playwright.chromium.launch(**launch_options)
        logging.info("Navegador configurado com sucesso")
        return browser
    except Exception as e:
        logging.error(f"Erro ao iniciar o browser: {e}")
        return None

def realizar_login(page, email, senha):
    """Realiza login no sistema Music Delivery"""
    try:
        logging.info(f"Iniciando processo de login para: {email}")
        page.goto("http://sistema.musicdelivery.com.br/login?login_error", wait_until="networkidle")
        page.select_option("select[name='tipo']", "client")
        page.fill("#login-username", email)
        page.fill("#login-password", senha)
        page.locator('button[type="submit"]').click()
        page.wait_for_load_state("networkidle")

        if "login_error" in page.url:
            logging.error(f"Falha ao efetuar Login para {email}! Verifique usuário/senha.")
            return None

        page.goto("http://sistema.musicdelivery.com.br/cliente/", wait_until="networkidle")

        try:
            cliente_nome = page.locator("div.col-sm-7 h2").text_content().strip()
        except Exception:
            cliente_nome = "Cliente não identificado"
            logging.warning("Nome do cliente não encontrado após login")

        logging.info(f"Login realizado com sucesso para o cliente: {cliente_nome}")
        return cliente_nome
    except Exception as e:
        logging.error(f"Erro durante o login para {email}: {e}")
        return None

def extrair_tabela(page):
    """Extrai dados da tabela de logs"""
    base_url = "http://sistema.musicdelivery.com.br/logs"
    offset = 0
    all_data = []
    page_count = 1
    MAX_PAGES_TO_CHECK = 100

    logging.info("Iniciando extração de dados da tabela de logs")

    while True:
        url_to_visit = f"{base_url}/{offset}" if offset > 0 else base_url
        logging.info(f"Navegando para URL: {url_to_visit} (Página {page_count})")

        try:
            page.goto(url_to_visit, wait_until="networkidle", timeout=60000)
            time_module.sleep(2)
            rows = page.locator("table.table-striped tbody tr").all()

            if not rows:
                logging.info(f"Nenhuma linha de dados encontrada na página {page_count}. Fim da extração.")
                break

            current_page_data = []
            for row in rows:
                cols = row.locator("td").all()
                if len(cols) >= 4:
                    loja = cols[1].inner_text().strip()
                    identificador = cols[2].inner_text().strip()
                    atualizacao = cols[3].inner_text().strip()
                    current_page_data.append({
                        "Loja": loja,
                        "Identificador": identificador,
                        "Atualizado em": atualizacao
                    })
                else:
                    logging.warning(f"Linha com menos colunas na página {page_count}. Pulando.")

            if not current_page_data:
                logging.info(f"Nenhum dado válido na página {page_count}. Fim da extração.")
                break

            all_data.extend(current_page_data)
            logging.info(f"{len(current_page_data)} lojas extraídas da página {page_count}")

            offset += 30
            page_count += 1

            if page_count > MAX_PAGES_TO_CHECK:
                logging.warning(f"Limite de {MAX_PAGES_TO_CHECK} páginas atingido. Parando extração.")
                break

        except Exception as e:
            logging.error(f"Erro ao visitar {url_to_visit}: {e}")
            break

    df = pd.DataFrame(all_data)
    logging.info(f"Extração concluída. Total de {len(df)} lojas coletadas")
    return df

def analisar_sincronizacao(df):
    """Analisa dados de sincronização das lojas"""
    tz_sp = ZoneInfo("America/Sao_Paulo")
    df['Data Atualizacao'] = pd.to_datetime(df['Atualizado em'], dayfirst=True).dt.tz_localize(tz_sp, ambiguous='NaT', nonexistent='shift_forward')
    
    now = datetime.now(tz_sp)
    dt_inicio = datetime.combine(now.date(), time.min).replace(tzinfo=tz_sp)
    dt_fim = datetime.combine(now.date(), time.max).replace(tzinfo=tz_sp)

    df['Sincronizada'] = df['Data Atualizacao'].between(dt_inicio, dt_fim)
    df['Tempo Atraso'] = df.apply(
        lambda row: dt_inicio - row['Data Atualizacao'] if not row['Sincronizada'] else timedelta(0),
        axis=1
    )

    total_lojas = int(len(df))
    sincronizadas = int(df['Sincronizada'].sum())
    atrasadas = int(total_lojas - sincronizadas)
    percentual_sincronizadas = float((sincronizadas / total_lojas * 100) if total_lojas else 0)
    percentual_atrasadas = float(100 - percentual_sincronizadas)

    resumo = {
        'total': total_lojas,
        'sincronizadas': sincronizadas,
        'atrasadas': atrasadas,
        'percentual_sincronizadas': percentual_sincronizadas,
        'percentual_atrasadas': percentual_atrasadas
    }

    return df, resumo

def salvar_excel_relatorio(df, resumo, cliente_nome):
    """Gera relatório em Excel (mantido para compatibilidade)"""
    arquivo_excel = f"relatorio_{cliente_nome.replace(' ', '_').replace('.', '').replace('/', '_')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Lojas"

    headers = ["Loja", "Identificador", "Atualizado em", "Sincronizada", "Tempo Atraso"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def format_timedelta(td):
        if td == timedelta(0):
            return ""
        total_seconds = int(td.total_seconds())
        dias = total_seconds // 86400
        horas = (total_seconds % 86400) // 3600
        minutos = (total_seconds % 3600) // 60
        return f"{dias}d {horas}h {minutos}m"

    for _, row in df.iterrows():
        ws.append([
            row['Loja'],
            row['Identificador'],
            row['Atualizado em'],
            "Sim" if row['Sincronizada'] else "Não",
            format_timedelta(row['Tempo Atraso'])
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[column].width = max_length + 2

    # Adicionar resumo
    start_row = ws.max_row + 2
    ws.cell(row=start_row, column=1, value="Resumo").font = Font(bold=True, size=12)
    ws.cell(row=start_row + 1, column=1, value="Total de Lojas:")
    ws.cell(row=start_row + 1, column=2, value=resumo['total'])
    ws.cell(row=start_row + 2, column=1, value="Lojas sincronizadas:")
    ws.cell(row=start_row + 2, column=2, value=f"{resumo['sincronizadas']} ({resumo['percentual_sincronizadas']:.2f}%)")
    ws.cell(row=start_row + 3, column=1, value="Lojas atrasadas:")
    ws.cell(row=start_row + 3, column=2, value=f"{resumo['atrasadas']} ({resumo['percentual_atrasadas']:.2f}%)")

    # Criar gráfico
    if resumo['total'] > 0:
        labels = ['Sincronizadas', 'Atrasadas']
        sizes = [resumo['sincronizadas'], resumo['atrasadas']]
        colors = ['#4CAF50', '#F44336']

        plt.figure(figsize=(4,4))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', colors=colors, startangle=140)
        plt.title("Proporção de Lojas Sincronizadas x Atrasadas")
        plt.tight_layout()

        img_bytes = io.BytesIO()
        plt.savefig(img_bytes, format='png')
        plt.close()
        img_bytes.seek(0)

        img = Img(img_bytes)
        img.anchor = f'A{start_row + 5}'
        ws.add_image(img)

    wb.save(arquivo_excel)
    logging.info(f"Arquivo Excel salvo: {arquivo_excel}")
    return arquivo_excel, resumo['total']

def enviar_arquivo_telegram(arquivo_excel, cliente_nome, total_lojas, chat_id_to_send, incluir_supabase_info=True):
    """Envia arquivo via Telegram com informações do Supabase"""
    try:
        if not TELEGRAM_BOT_TOKEN:
            logging.error("TELEGRAM_BOT_TOKEN não configurado")
            return

        url_send_document = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
        url_send_message = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"

        emoji = "🏪" if total_lojas > 0 else "📜"
        hora_sp = datetime.now(ZoneInfo("America/Sao_Paulo"))
        origem = "GitHub Actions" if os.getenv("GITHUB_ACTIONS") else "Execução Local"

        mensagem_base = (
            f"{emoji} **Relatório de Lojas Music Delivery**\n"
            f"👤 **Cliente:** {cliente_nome}\n"
            f"📊 **Total de Lojas:** {total_lojas}\n"
            f"📅 **Data da Extração:** {hora_sp.strftime('%d/%m/%Y às %H:%M:%S')}\n"
            f"💻 **Origem:** {origem}"
        )
        
        if incluir_supabase_info:
            mensagem_base += (
                f"\n\n📈 **Dashboard Web:** Os dados foram salvos no banco de dados e estão disponíveis em tempo real no dashboard web!\n"
                f"🔗 **Acesse:** [Link do seu dashboard aqui]"
            )

        response_msg = requests.post(url_send_message, data={
            "chat_id": chat_id_to_send,
            "text": mensagem_base,
            "parse_mode": "Markdown"
        }, timeout=30)

        if response_msg.status_code != 200:
            logging.error(f"Erro ao enviar mensagem: {response_msg.text}")

        if arquivo_excel and os.path.exists(arquivo_excel):
            with open(arquivo_excel, "rb") as file:
                response_doc = requests.post(
                    url_send_document,
                    data={"chat_id": chat_id_to_send},
                    files={"document": file},
                    timeout=60
                )

            if response_doc.status_code == 200:
                logging.info(f"Arquivo enviado com sucesso via Telegram para {cliente_nome}")
            else:
                logging.error(f"Erro ao enviar arquivo para {cliente_nome}: {response_doc.text}")
        else:
            logging.error(f"Arquivo Excel não encontrado para envio - {cliente_nome}")
    except Exception as e:
        logging.error(f"Erro ao enviar via Telegram para {cliente_nome}: {e}")

def enviar_notificacao_erro(erro_msg, chat_id_to_send, cliente_nome="Sistema"):
    """Envia notificação de erro via Telegram"""
    try:
        if not TELEGRAM_BOT_TOKEN:
            return

        url_send_message = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        hora_sp = datetime.now(ZoneInfo("America/Sao_Paulo"))
        origem = "GitHub Actions" if os.getenv("GITHUB_ACTIONS") else "Execução Local"

        mensagem = (
            f"🚨 **Erro no Monitoramento de Lojas**\n"
            f"👤 **Cliente:** {cliente_nome}\n"
            f"❌ **Erro:** {erro_msg}\n"
            f"🕐 **Timestamp:** {hora_sp.strftime('%d/%m/%Y às %H:%M:%S')}\n"
            f"💻 **Origem:** {origem}"
        )

        requests.post(url_send_message, data={
            "chat_id": chat_id_to_send,
            "text": mensagem,
            "parse_mode": "Markdown"
        }, timeout=30)
    except Exception as e:
        logging.error(f"Erro ao enviar notificação: {e}")

def processar_cliente(browser, cliente_info):
    """Processa um cliente específico com integração Supabase"""
    context = browser.new_context()
    page = context.new_page()
    
    cliente_nome = cliente_info.get('nome', 'Cliente não identificado')
    email = cliente_info.get('email')
    senha = cliente_info.get('senha')
    chat_id = cliente_info.get('chat_id')
    
    logging.info(f"🔄 Processando cliente: {cliente_nome}")
    
    # Inicializar Supabase
    supabase = init_supabase()
    if not supabase:
        logging.error("Falha ao conectar com Supabase - processo abortado")
        return False
    
    # Criar registro de execução
    execucao_id = criar_execucao(supabase, cliente_info)
    if not execucao_id:
        logging.error(f"Falha ao criar execução para {cliente_nome}")
        return False
    
    try:
        # Realizar login
        nome_logado = realizar_login(page, email, senha)
        if not nome_logado:
            erro_msg = f"Falha no login - verifique credenciais para {cliente_nome}"
            logging.error(erro_msg)
            finalizar_execucao(supabase, execucao_id, {}, "erro", erro_msg)
            log_execucao(cliente_nome, "erro", erro_msg)
            enviar_notificacao_erro(erro_msg, chat_id, cliente_nome)
            return False

        # Extrair dados
        df = extrair_tabela(page)

        if df.empty:
            logging.info(f"Nenhuma loja encontrada para {cliente_nome}")
            resumo = {'total': 0, 'sincronizadas': 0, 'atrasadas': 0, 'percentual_sincronizadas': 0, 'percentual_atrasadas': 0}
            finalizar_execucao(supabase, execucao_id, resumo, "sem_dados", "Nenhuma loja encontrada")
            log_execucao(cliente_nome, "sem_dados", "Nenhuma loja encontrada")
            
            mensagem = f"📌 **Cliente:** {cliente_nome}\n❌ Nenhuma loja encontrada no sistema."
            if TELEGRAM_BOT_TOKEN:
                requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage", data={
                    "chat_id": chat_id,
                    "text": mensagem,
                    "parse_mode": "Markdown"
                })
            return True

        # Análise de sincronização
        df, resumo = analisar_sincronizacao(df)

        # Salvar dados no Supabase
        sucesso_lojas = salvar_dados_lojas_supabase(supabase, execucao_id, df, cliente_info)
        if not sucesso_lojas:
            logging.error(f"Falha ao salvar dados das lojas para {cliente_nome}")

        # Finalizar execução
        finalizar_execucao(supabase, execucao_id, resumo, "sucesso")
        
        # Atualizar métricas periódicas
        atualizar_metricas_periodicas(supabase, cliente_info, resumo)

        # Gerar relatório Excel (opcional, para compatibilidade)
        arquivo_excel = None
        if os.getenv("GERAR_EXCEL", "true").lower() == "true":
            arquivo_excel, total_lojas = salvar_excel_relatorio(df, resumo, cliente_nome)
        else:
            total_lojas = resumo['total']

        # Log de sucesso
        log_execucao(cliente_nome, "sucesso", f"Dados salvos no Supabase - {total_lojas} lojas", total_lojas)
        
        # Enviar notificação via Telegram
        if arquivo_excel:
            enviar_arquivo_telegram(arquivo_excel, cliente_nome, total_lojas, chat_id, True)
            
            # Remover arquivo após envio (no GitHub Actions)
            if os.getenv("GITHUB_ACTIONS"):
                try:
                    os.remove(arquivo_excel)
                    logging.info(f"Arquivo {arquivo_excel} removido após envio")
                except Exception:
                    logging.warning(f"Não foi possível remover arquivo {arquivo_excel}")
        else:
            # Enviar apenas notificação de sucesso sem arquivo
            enviar_notificacao_sucesso_supabase(cliente_nome, resumo, chat_id)

        logging.info(f"✅ Cliente {cliente_nome} processado com sucesso! Dados salvos no Supabase.")
        return True

    except Exception as e:
        erro_msg = f"Erro crítico ao processar {cliente_nome}: {str(e)}"
        logging.critical(erro_msg)
        finalizar_execucao(supabase, execucao_id, {}, "erro", erro_msg)
        log_execucao(cliente_nome, "erro", erro_msg)
        enviar_notificacao_erro(erro_msg, chat_id, cliente_nome)
        return False
    finally:
        if context:
            context.close()

def enviar_notificacao_sucesso_supabase(cliente_nome, resumo, chat_id):
    """Envia notificação de sucesso com dados do Supabase"""
    try:
        if not TELEGRAM_BOT_TOKEN:
            return

        url_send_message = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        hora_sp = datetime.now(ZoneInfo("America/Sao_Paulo"))
        origem = "GitHub Actions" if os.getenv("GITHUB_ACTIONS") else "Execução Local"

        mensagem = (
            f"✅ **Dados Coletados e Salvos**\n"
            f"👤 **Cliente:** {cliente_nome}\n"
            f"📊 **Total de Lojas:** {resumo['total']}\n"
            f"🟢 **Sincronizadas:** {resumo['sincronizadas']} ({resumo['percentual_sincronizadas']:.1f}%)\n"
            f"🔴 **Atrasadas:** {resumo['atrasadas']} ({resumo['percentual_atrasadas']:.1f}%)\n"
            f"📅 **Coletado em:** {hora_sp.strftime('%d/%m/%Y às %H:%M:%S')}\n"
            f"💾 **Status:** Dados salvos no banco de dados\n"
            f"📈 **Dashboard:** Disponível no painel web\n"
            f"💻 **Origem:** {origem}"
        )

        requests.post(url_send_message, data={
            "chat_id": chat_id,
            "text": mensagem,
            "parse_mode": "Markdown"
        }, timeout=30)
        
    except Exception as e:
        logging.error(f"Erro ao enviar notificação de sucesso: {e}")

def obter_estatisticas_supabase(cliente_nome=None, dias=30):
    """Função para obter estatísticas dos dados no Supabase (para uso em dashboards)"""
    try:
        supabase = init_supabase()
        if not supabase:
            return None
        
        # Data de referência
        data_limite = datetime.now(ZoneInfo("America/Sao_Paulo")) - timedelta(days=dias)
        
        # Query base
        query = supabase.table('execucoes').select('*').gte('executado_em', data_limite.isoformat())
        
        if cliente_nome:
            query = query.eq('cliente_nome', cliente_nome)
        
        response = query.order('executado_em', desc=True).execute()
        
        if response.data:
            df = pd.DataFrame(response.data)
            
            # Processar dados para o dashboard
            df['executado_em'] = pd.to_datetime(df['executado_em'])
            df_agrupado = df.groupby(df['executado_em'].dt.date).agg(
                total_lojas=('total_lojas', 'max'),
                lojas_sincronizadas=('lojas_sincronizadas', 'max'),
                lojas_atrasadas=('lojas_atrasadas', 'max'),
                percentual_sincronizadas=('percentual_sincronizadas', 'mean')
            ).reset_index()
            
            return df_agrupado.to_dict('records')
        else:
            return []
        
    except Exception as e:
        logging.error(f"Erro ao obter estatísticas do Supabase: {e}")
        return None

def main():
    """Função principal"""
    logging.info("Iniciando monitoramento de clientes")
    browser = None
    total_processados = 0
    total_sucessos = 0
    
    try:
        clientes = carregar_base_clientes()
        if not clientes:
            logging.error("Não há clientes para processar. Encerrando.")
            return
        
        browser = setup_browser(headless=os.getenv("GITHUB_ACTIONS") is not None)
        if not browser:
            logging.critical("Falha ao iniciar o navegador. Encerrando.")
            return

        for cliente in clientes:
            total_processados += 1
            if processar_cliente(browser, cliente):
                total_sucessos += 1
            
    except Exception as e:
        logging.critical(f"Erro crítico na execução principal: {e}")
        enviar_notificacao_erro(f"Erro crítico: {e}", ADMIN_CHAT_ID)
    finally:
        if browser:
            browser.close()
            logging.info("Navegador fechado")
        
    logging.info(f"🎯 Processamento finalizado: {total_sucessos} sucessos, {total_processados - total_sucessos} falhas")
    if total_sucessos == 0 and total_processados > 0:
        enviar_notificacao_erro("Nenhum cliente foi processado com sucesso.", ADMIN_CHAT_ID, "Sistema")

if __name__ == "__main__":
    main()